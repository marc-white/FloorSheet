# Module that creates the scoresheet

# import csv
import re
import pdfrw
import openpyxl
from openpyxl.styles import Color, PatternFill

# import subprocess

from django.conf import settings

# from django.core.files.storage import DefaultStorage

# from fdfgen import forge_fdf

ANNOT_KEY = "/Annots"
ANNOT_FIELD_KEY = "/T"
ANNOT_VAL_KEY = "/V"
ANNOT_RECT_KEY = "/Rect"
SUBTYPE_KEY = "/Subtype"
WIDGET_SUBTYPE_KEY = "/Widget"

STAFF_ROLE_KEY = "sSmM"
CAPTAIN_ROLE_KEY = "cC"
GOALIE_ROLE_KEY = "gG"

MAX_NO_OF_PLAYERS = 20
MAX_NO_OF_STAFF = 5

XLS_COLUMN_GC = "R"
XLS_COLUMN_PLAYER_NO = "S"
XLS_COLUMN_PLAYER_NAME = "T"
XLS_COLUMN_DOB = "Y"

XLS_ROW_HOME_FIRST = 4
XLS_ROW_AWAY_FIRST = 28

XLS_HOME_TEAM_NAME = "U2"
XLS_HOME_STAFF_1 = "W24"
XLS_HOME_STAFF_2 = "AB24"
XLS_HOME_STAFF_3 = "AH24"
XLS_HOME_STAFF_4 = "AB25"
XLS_HOME_STAFF_5 = "AH25"

XLS_AWAY_TEAM_NAME = "U26"
XLS_AWAY_STAFF_1 = "W48"
XLS_AWAY_STAFF_2 = "AB48"
XLS_AWAY_STAFF_3 = "AH48"
XLS_AWAY_STAFF_4 = "AB49"
XLS_AWAY_STAFF_5 = "AH49"

XLS_CELL_ASSOCIATION = "E6"
XLS_CELL_COMPETITION = "E8"
XLS_CELL_DIVISION = "P8"
XLS_CELL_VENUE = "E10"
XLS_CELL_DATE = "D12"
XLS_CELL_MATCH_ID = "I13"
XLS_CELL_START_TIME = "D14"
XLS_CELL_END_TIME = "L14"
XLS_CELL_REFEREE_PAIR = "A38"
XLS_CELL_SECRETARY_NAME = "A34"

XLS_COLOR_JUNIOR = Color("FFFF99")
XLS_FILL_JUNIOR = PatternFill("solid", fgColor=XLS_COLOR_JUNIOR)
XLS_AGE_LIMIT_JUNIOR = 18

XLS_COLOR_BORROWED = Color("CCFFCC")
XLS_FILL_BORROWED = PatternFill("solid", fgColor=XLS_COLOR_BORROWED)
XLS_BORROWED_MARKER = "(B)"

XLS_COLOR_MEMBERSHIP = Color("CCFFFF")
XLS_FILL_MEMBERSHIP = PatternFill("solid", fgColor=XLS_COLOR_MEMBERSHIP)
XLS_MEMBERSHIP_MARKER = "(MEMB)"


DATE_DISPLAY_FORMAT = "%d %b %Y"
TIME_DISPLAY_FORMAT = "%H:%M"
DATE_DISPLAY_FORMAT_PLAYER = "%d/%m/%y"
DATE_DISPLAY_REGEX_PLAYER_AGE = re.compile(
    r"^\s*(?P<day>[0-9]{2})/(?P<month>[0-9]{2})/(?P<year>[0-9]{2})\s*\((?P<age>[0-9]*)\)\s*$"
)


def try_int_or_zero(x):
    try:
        return int(x)
    except ValueError:
        return 0


def compute_player_age(dob, ref_date):
    """
    Computes an player age.

    This is an inefficient recursive algorithm, but is the easiest
    way to handle leap years, etc.
    """
    ref_date = ref_date.date()  # Game time is stored as datetime
    assert ref_date > dob, "Cannot compute an age for a future birthday!"

    age = 0
    while ref_date.replace(year=ref_date.year - age - 1) >= dob:
        age += 1
    return age


def parse_team_list(team_obj, game_date=None):
    """
    Parse a team player list into a dict to be written to PDF
    """
    players = team_obj.player_set.filter(is_staff=False).order_by("number", "name")
    staff = team_obj.player_set.filter(is_staff=True).order_by("number", "name")
    # No-op if no players on team
    if players.count() == 0 and staff.count() == 0:
        return {}

    # Build and write back a player list and staff list
    data_dict = {}
    for i, player in enumerate(players[:20]):
        no = "{:02d}".format(i + 1)
        data_dict["PlayerGC{}".format(no)] = ""
        if player.is_captain:
            data_dict["PlayerGC{}".format(no)] += "C"
        if player.is_goalie:
            data_dict["PlayerGC{}".format(no)] += "G"
        if player.number is not None:
            data_dict["PlayerNo{}".format(no)] = str(player.number) or ""
        else:
            data_dict["PlayerNo{}".format(no)] = ""
        data_dict["PlayerName{}".format(no)] = player.name
        data_dict["PlayerDOB{}".format(no)] = (
            player.dob.strftime(DATE_DISPLAY_FORMAT_PLAYER) if player.dob else ""
        )
        if game_date is not None and player.dob is not None:
            data_dict[
                "PlayerDOB{}".format(no)
            ] += f" ({compute_player_age(player.dob, game_date)})"

    for i, player in enumerate(staff):
        no = "{:1d}".format(i + 1)
        data_dict["Off{}".format(no)] = player.name

    # import pdb; pdb.set_trace()
    return data_dict


def create_scoresheet_data(*args, **kwargs):
    """
    Start filling out a floorball scoresheet, based on a passed info

    Returns
    -------
    pdfrw.PdfFileReader
        A PdfFileReader instance. This can be written to disk by sending
        it to a PdfFileWriter instance.
    """

    # Build a data-dict from the input kwargs
    data_dict = {}

    # print(kwargs)

    # Single input things
    if kwargs.get("assoc"):
        pass
    if kwargs.get("comp"):
        # data_dict['Competition'] = "{} {}".format(
        #     kwargs['comp'].comp.comp,
        #     kwargs['comp'].get_div_display(),
        # )
        data_dict["comp_obj"] = kwargs["comp"]
        data_dict["Association"] = kwargs["comp"].comp.assoc
    if kwargs.get("venue"):
        data_dict["Venue"] = kwargs["venue"]
    if kwargs.get("match_id"):
        data_dict["MatchNo"] = kwargs["match_id"]
    if kwargs.get("start_time"):
        data_dict["Date"] = kwargs["start_time"].strftime(DATE_DISPLAY_FORMAT)
        data_dict["StartTime"] = kwargs["start_time"].strftime(TIME_DISPLAY_FORMAT)
    if kwargs.get("duty_team"):
        data_dict["MatchSecName"] = "[{}{}]".format(
            kwargs["duty_team"].__str__(),
            " ({})".format(kwargs["duty_team"].color.capitalize())
            if kwargs["duty_team"].color
            else "",
        )
    # print('- create_scoresheet: Added basic team info')

    for team in ["home", "away"]:
        team_field_code = team.capitalize()
        team_obj = kwargs.get("{}_team".format(team))
        if team_obj is None:
            continue

        # Write in the team name
        data_dict["{}TeamName".format(team_field_code)] = "{}{}".format(
            team_obj.team_name,
            " ({})".format(team_obj.color.capitalize()) if team_obj.color else "",
        )

        # Add in the player/staff list
        data_dict.update(
            {
                "{}{}".format(team_field_code, k): v
                for k, v in list(
                    parse_team_list(team_obj, game_date=kwargs["start_time"]).items()
                )
            }
        )
    # print('- create_scoresheet: Added team lists')
    # print('- create_scoresheet: Final data_dict is...\n{}\n'.format(data_dict))

    # import pdb; pdb.set_trace()
    return data_dict


def create_scoresheet_pdf(template, *args, **kwargs):
    data_dict = create_scoresheet_data(*args, **kwargs)

    # Do the necessary manipulation set up the competition name
    data_dict["Competition"] = "{} {}".format(
        data_dict["comp_obj"].comp.comp,
        data_dict["comp_obj"].get_div_display(),
    )
    del data_dict["comp_obj"]

    # Open up the template form
    # print("Rendering PDF...")
    template_pdf = pdfrw.PdfReader(template)
    template_pdf.Root.AcroForm.update(
        pdfrw.PdfDict(NeedAppearances=pdfrw.PdfObject("true"))
    )
    annotations = template_pdf.pages[0][ANNOT_KEY]
    for annotation in annotations:
        if annotation[SUBTYPE_KEY] == WIDGET_SUBTYPE_KEY:
            if annotation[ANNOT_FIELD_KEY]:
                key = annotation[ANNOT_FIELD_KEY][1:-1]
                if key in list(data_dict.keys()):
                    # print('- create_scoresheet: Adding key {}'.format(key))
                    annotation.update(pdfrw.PdfDict(V="{}".format(data_dict[key])))

    # print("Rendering complete!")
    return template_pdf


def create_scoresheet_xlsx(template, *args, **kwargs):
    data_dict = create_scoresheet_data(*args, **kwargs)

    # Do the necessary manipulation set up the competition name
    data_dict["Competition"] = data_dict["comp_obj"].comp.comp
    # data_dict['Division'] = data_dict['comp_obj'].div.upper()
    data_dict["Division"] = "".join(
        [_.upper()[0] for _ in data_dict["comp_obj"].get_div_display().split()]
    )
    del data_dict["comp_obj"]

    # Open up the template
    template_xlsx = openpyxl.reader.excel.load_workbook(template)
    template_record = template_xlsx.get_sheet_by_name("Record")

    # Set up the game parameters
    for target_cell, target_data in {
        XLS_CELL_ASSOCIATION: "Association",
        XLS_CELL_COMPETITION: "Competition",
        XLS_CELL_DIVISION: "Division",
        XLS_CELL_VENUE: "Venue",
        XLS_CELL_DATE: "Date",
        XLS_CELL_MATCH_ID: "MatchNo",
        XLS_CELL_START_TIME: "StartTime",
        XLS_CELL_SECRETARY_NAME: "MatchSecName",
    }.items():
        try:
            template_record[target_cell] = data_dict.get(target_data, None)
        except ValueError:
            template_record[target_cell] = data_dict.get(target_data).__str__()

    for team in ["Home", "Away"]:
        if data_dict.get(f"{team}TeamName") is not None:

            template_record[globals()[f"XLS_{team.upper()}_TEAM_NAME"]] = data_dict.get(
                f"{team}TeamName"
            )

            for i in range(0, MAX_NO_OF_PLAYERS):
                istr = f"{i+1:02d}"
                if data_dict.get(f"{team}PlayerName{istr}") is not None:
                    row = globals()[f"XLS_ROW_{team.upper()}_FIRST"] + i
                    template_record[f"{XLS_COLUMN_GC}{row}"] = data_dict.get(
                        f"{team}PlayerGC{istr}"
                    )
                    template_record[f"{XLS_COLUMN_PLAYER_NO}{row}"] = data_dict.get(
                        f"{team}PlayerNo{istr}"
                    )
                    template_record[f"{XLS_COLUMN_PLAYER_NAME}{row}"] = data_dict.get(
                        f"{team}PlayerName{istr}"
                    )
                    template_record[f"{XLS_COLUMN_DOB}{row}"] = data_dict.get(
                        f"{team}PlayerDOB{istr}"
                    )
                    age_match = re.match(
                        DATE_DISPLAY_REGEX_PLAYER_AGE,
                        data_dict.get(f"{team}PlayerDOB{istr}"),
                    )
                    if age_match and int(age_match.group("age")) < XLS_AGE_LIMIT_JUNIOR:
                        template_record[f"{XLS_COLUMN_DOB}{row}"].fill = XLS_FILL_JUNIOR
                    if XLS_BORROWED_MARKER in data_dict.get(f"{team}PlayerName{istr}"):
                        template_record[
                            f"{XLS_COLUMN_PLAYER_NAME}{row}"
                        ].fill = XLS_FILL_BORROWED
                    if XLS_MEMBERSHIP_MARKER in data_dict.get(
                        f"{team}PlayerName{istr}"
                    ):
                        template_record[
                            f"{XLS_COLUMN_PLAYER_NAME}{row}"
                        ].fill = XLS_FILL_MEMBERSHIP

            for i in range(0, MAX_NO_OF_STAFF):
                istr = f"{i+1:01d}"
                template_record[
                    globals()[f"XLS_{team.upper()}_STAFF_{istr}"]
                ] = data_dict.get(f"{team}Off{istr}")

    return template_xlsx
