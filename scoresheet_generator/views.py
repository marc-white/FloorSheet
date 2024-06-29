# -*- coding: utf-8 -*-

import os
import datetime
import subprocess

from django.shortcuts import render
import openpyxl.drawing
from . import forms, writer, create_pdf_from_xlsx
from django.forms import formset_factory
from django.contrib.staticfiles import finders
# from django.core.files.base import ContentFile
# from django.core.files.storage import default_storage
from django.shortcuts import redirect
from django.conf import settings

from pdfrw import PdfWriter, IndirectPdfDict, PdfReader, PdfObject, PdfDict
import openpyxl
from openpyxl_image_loader import SheetImageLoader
SHEET_IMAGE_LOCATIONS = {
    "Record": ["A2", ],
    "Completion": ["A1", ], 
}
SHEET_IMAGE_SIZES = {
    "Record": {
        "width": 67-2,
        "height": 56-2,
    },
    "Completion": {
        "width": 82-2,
        "height": 76-2,
    }
}
CONVERTED_PDF_SIZE_LIMIT = 250

from django.conf import settings

# Create your views here.


def generate(request):
    
    ss_formset = formset_factory(forms.ScoresheetGeneratorForm, extra=5)

    if request.method == 'POST':

        fmt = request.POST.get("create")

        convert_xlsx = False
        if fmt == "xlsx-pdf":
            convert_xlsx = True
            fmt = "xlsx"

        if fmt == "pdf":
            template_path = finders.find('scoresheet_generator/pdf/'
                                        'IFF-Match-Record-2022-Temp-Form-MVP3'
                                        '.pdf')
            writer_func = writer.create_scoresheet_pdf
        elif fmt == "xlsx":
            template_path = finders.find('scoresheet_generator/xlsx/'
                                        'IFFMatchRecord2022_blank.xlsx')
            writer_func = writer.create_scoresheet_xlsx
        else:
            raise ValueError(f"Format {fmt} is unknown")

        ss_formset = ss_formset(request.POST)
        if ss_formset.is_valid():
            sheets = []
            start_date = None
            end_date = None
            for form in ss_formset:
                if form.is_valid() and form.has_changed():
                    template = writer_func(
                        template_path,
                        comp=form.cleaned_data.get('comp'),
                        start_time=form.cleaned_data.get('start_time'),
                        match_id=form.cleaned_data.get('match_id'),
                        home_team=form.cleaned_data.get('home_team'),
                        away_team=form.cleaned_data.get('away_team'),
                        duty_team=form.cleaned_data.get('duty_team'),
                        venue=form.cleaned_data.get('venue'),
                    )
                    if start_date is None or start_date > form.cleaned_data.get('start_time'):
                        start_date = form.cleaned_data.get('start_time')
                    if end_date is None or end_date < form.cleaned_data.get('start_time'):
                        end_date = form.cleaned_data.get('start_time')
                    # p = os.path.join(
                    #     settings.MEDIA_ROOT,
                    #     '{}{}_{}-v-{}.{}'.format(
                    #         form.cleaned_data.get('start_time').strftime(
                    #             '%Y-%m-%d_%H%M'),
                    #         '_{}'.format(form.cleaned_data.get('match_id')) if
                    #         form.cleaned_data.get('match_id') else
                    #         '',
                    #         form.cleaned_data.get('home_team').team_name[:4],
                    #         form.cleaned_data.get('away_team').team_name[:4],
                    #         fmt
                    #     ))
                    # print(f"Created filename {p}")

                    sheets.append(template)

            # Merge the individual sheets into a master file
            # (easier file handling, and uses less API hits for 
            #  converting XSLX->PDF via Adobe APIs)
            if start_date is None or end_date is None:
                output_fn = f"sheets_generated_at_{datetime.datetime.now().strftime('%Y-%m-%d-%H%M')}.{fmt}"
            elif start_date.date() == end_date.date():
                output_fn = f"sheets_{start_date.strftime('%Y-%m-%d')}.{fmt}"
            else:
                output_fn = f"sheets_{start_date.strftime('%Y-%m-%d')}-TO-{end_date.strftime('%Y-%m-%d')}.{fmt}"
            output_fn_full = os.path.join(settings.MEDIA_ROOT, output_fn)

            if fmt == "pdf":
                # import pdb; pdb.set_trace()
                output_pdf = PdfWriter()
                for sheet in sheets:
                    for page in sheet.pages:
                        output_pdf.addpage(page)
                output_pdf.write(output_fn_full)
                # print("Attempting redirect to %s", '/'.join(output_fn_full.split(os.sep)[-2:]))
            
            if fmt == "xlsx":
                output_wb = openpyxl.reader.excel.load_workbook(template_path)
                sheet_names_orig = output_wb.get_sheet_names()
                # Make duplicate match record sheets as required
                for i in range(1, len(sheets)):
                    for sheet_name in sheet_names_orig:
                        new_sheet = output_wb.copy_worksheet(output_wb[sheet_name])
                        new_sheet.title = f"{sheet_name}{i+1}"

                # Iterate over the match records in memory, and copy them
                # in to the output workbook
                for i, sheet in enumerate(sheets):
                    for sheet_name in sheet_names_orig:
                        filled_sheet = sheet[sheet_name]
                        blank_sheet = output_wb[
                            f"{sheet_name}{i+1}" if i != 0 else sheet_name
                            ]
                        mr = filled_sheet.max_row
                        mc = filled_sheet.max_column

                        for r in range(1, mr+1):
                            for c in range(1, mc+1):
                                try:
                                    blank_sheet.cell(
                                        row=r, column=c
                                        ).value = filled_sheet.cell(row=r, column=c).value
                                except AttributeError:  # Merged cell, ignore
                                    pass
                        if i != 0:
                            loader = SheetImageLoader(filled_sheet)
                            for im_pos in SHEET_IMAGE_LOCATIONS[sheet_name]:
                                im = openpyxl.drawing.image.Image(loader.get(im_pos))
                                im.height = SHEET_IMAGE_SIZES[sheet_name]["height"]
                                im.width = SHEET_IMAGE_SIZES[sheet_name]["width"]
                                blank_sheet.add_image(im, im_pos)
                
                output_wb.save(output_fn_full)

                if convert_xlsx:
                    output_fn_full = create_pdf_from_xlsx.create_pdf_from_xlsx(output_fn_full)

                    # The Adobe API sometimes includes blank pages in its return. We can 
                    # identify and eliminate these by looking at
                    # page.Contents["/Length"] and deleting pages below a certain limit
                    converted_pdf = PdfReader(output_fn_full)
                    final_pdf = PdfWriter()
                    for page in converted_pdf.pages:
                        if int(page.Contents["/Length"]) > CONVERTED_PDF_SIZE_LIMIT:
                            final_pdf.addpage(page)
                    final_pdf.write(output_fn_full)


            return redirect('/'.join(output_fn_full.split(os.sep)[-2:]))


            # Make and serve a zip of these
            # zip_name = 'sheets_made_at_' + datetime.datetime.now().strftime('%Y-%m-%d-%H%M') + '.zip'
            # print(f"Preparing to zip to {zip_name}")
            # zip_path = os.path.join(settings.MEDIA_ROOT, zip_name)
            # subprocess.check_call(['zip', '-jFS', zip_path] + pdfs)
            # print("Zip complete!")
            # return redirect('/'.join(zip_path.split(os.sep)[-2:]))

    return render(request, 'scoresheet_generator/generator.html',
                  {
                      'ss_formset': ss_formset,
                  })

